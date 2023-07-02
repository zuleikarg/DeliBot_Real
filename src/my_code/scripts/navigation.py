#!/usr/bin/env python3

#LIBRARIES
import rospy
import actionlib
import openpyxl
import requests
import os

from my_code.msg import data_employee, data_recognition

from geometry_msgs.msg import Twist
from tf2_msgs.msg import TFMessage
from nav_msgs.msg import Odometry
from move_base_msgs.msg import MoveBaseAction, MoveBaseGoal
from geometry_msgs.msg import Pose
from tf.transformations import euler_from_quaternion, quaternion_from_euler
from kobuki_msgs.msg import Sound, BumperEvent

# print(cv2.__version__)
import numpy as np

tam = 0
st =False
max_count = 25

#INITIAL POSE - MODIFICABLE
ini_pos =[-1.0,0.5,0.0]

class MoveRobot:
    def __init__(self):
        self.folder = os.environ.get('DELIBOT_PATH', '')

        self.success = False
        self.start = False
        self.pos_x = []
        self.pos_y = []
        self.ang_z = []

        self.prev_pos_x = 0
        self.prev_pos_y = 0
        self.prev_ang_z = 0

        self.curr_pos_x = 0
        self.curr_pos_y = 0
        self.curr_ang_z = 0

        self.recognition = 0
        self.id_tele =""

        self.message_s = Sound()
        self.count = 0

        self.spinning = False                                                                       #To make it work when spins
        self.collected = False

        rospy.Subscriber('/odom', Odometry, self.odom) 
        self.move_pub = rospy.Publisher('/cmd_vel_mux/input/navi', Twist, queue_size=10) #CAMBIADO------------------------------------------
        self.speech_pub = rospy.Publisher('/mobile_base/commands/sound', Sound, queue_size=10)

    def collect(self,data):
        self.collected = True


    def odom(self, data):
        self.curr_pos_x = data.pose.pose.position.x 
        self.curr_pos_y = data.pose.pose.position.y
        orientation_quat = [data.pose.pose.orientation.x,data.pose.pose.orientation.y,data.pose.pose.orientation.z,data.pose.pose.orientation.w]
        (ang_x, ang_y, self.curr_ang_z) = euler_from_quaternion(orientation_quat) 

    def get_recog(self,data):
        if(data.recog == True and self.spinning == True):
            self.recognition = self.recognition +1
        else:
            if(self.recognition<3):
                self.recognition = 0


    def callback(self,data):
        global tam, st
        # Define variable to load the dataframe
        dataframe = openpyxl.load_workbook(self.folder + "/tfg_ros_real/src/my_code/datos_empleados.xlsx")
        # Define variable to read sheet
        dataframe1 = dataframe.active
        
        # Iterate the loop to read the cell values
        for row in dataframe1.iter_rows(1,dataframe1.max_row):
            if(row[0].value == data.name):
                self.id_tele = row[3].value
                break
        
        aux = 4
        raw_pos = row[aux].value.split("_")

        for j in range (4, len(row)):
            raw_pos = row[aux].value.split("_")

            self.pos_x.append(raw_pos[0])
            self.pos_y.append(raw_pos[1])
            self.ang_z.append(raw_pos[2])
            # print("Hola")
            aux = aux +1
        tam = len(self.pos_x)

        st = True
        self.start = True


        # Iterate the loop to read the cell values

    def spin_move(self,i):
        global max_count
        # print("SPIN")
        self.spinning = True

        self.prev_pos_x = self.curr_pos_x
        self.prev_pos_y = self.curr_pos_y
        self.prev_ang_z = self.curr_ang_z

        message = Twist()

        angle = 0.0

        message.linear.x = 0.0
        message.linear.y = 0.0
        message.linear.z = 0.0
        message.angular.x = 0.0
        message.angular.y = 0.0
        message.angular.z = 0.0
    

        while(angle < 3.3 and not rospy.is_shutdown() and self.collected == False):
            if(self.curr_ang_z < 0):
                self.curr_ang_z = self.curr_ang_z * -1

            if(self.prev_ang_z<0):
                self.prev_ang_z = self.prev_ang_z*-1

            angle = angle + abs(self.curr_ang_z - self.prev_ang_z)

            self.prev_ang_z = self.curr_ang_z

            # print(f'Antes:{self.prev_ang_z}, Actual: {self.curr_ang_z}, Angle: {angle}')

            if(self.recognition < 3):
                # print(self.recognition)
                message.angular.z = 0.5
                self.message_s.value = 2
            else:
                message.angular.z = 0.0
                self.message_s.value = 3
                max_count = 50


            if(self.count >= max_count):
                self.speech_pub.publish(self.message_s)
                self.count = 0
            else:
                self.count = self.count +1

            self.move_pub.publish(message)
            rate = rospy.Rate(10) # 10hz
            rate.sleep()
        
        self.spinning = False
        message.angular.z = 0.0
        self.move_pub.publish(message)



    #Execute the trayectories to the points of the Arucos saved
    def movebase_client(self,i):

        if(self.start == True):
            # print(i)
            # Create an action client called "move_base" with action definition file "MoveBaseAction"
            client = actionlib.SimpleActionClient('move_base',MoveBaseAction)
                
            # Waits until the action server has started up and started listening for goals.
            client.wait_for_server()

            # Creates a new goal with the MoveBaseGoal constructor
            goal = MoveBaseGoal()
            goal.target_pose.header.frame_id = "map"
            goal.target_pose.header.stamp = rospy.Time.now()
            # Move 0.5 meters forward along the axis of the "map" coordinate frame 
            if(i == -1):
                goal.target_pose.pose.position.x = float(ini_pos[0])
                goal.target_pose.pose.position.y = float(ini_pos[1])

                goal.target_pose.pose.orientation.x = 0.0
                goal.target_pose.pose.orientation.y = 0.0
                goal.target_pose.pose.orientation.z = float(ini_pos[2])
                goal.target_pose.pose.orientation.w = 1.0

            else:
                goal.target_pose.pose.position.x = float(self.pos_x[i])
                goal.target_pose.pose.position.y = float(self.pos_y[i])

                goal.target_pose.pose.orientation.x = 0.0
                goal.target_pose.pose.orientation.y = 0.0
                goal.target_pose.pose.orientation.z = float(self.ang_z[i])
                goal.target_pose.pose.orientation.w = 1.0

            # No rotation of the mobile base frame w.r.t. map frame
            # goal.target_pose.pose.orientation.w = 1.0

            # Sends the goal to the action server.
            client.send_goal(goal)
            # Waits for the server to finish performing the action.
            wait = client.wait_for_result()
            # If the result doesn't arrive, assume the Server is not available

            if not wait:
                rospy.logerr("Action server not available!")
                rospy.signal_shutdown("Action server not available!")
            else:
                # Result of executing the action
                return client.get_result()    



#MAIN
if __name__ == '__main__':
    rospy.init_node('navigation_node', anonymous=True)

    nav = MoveRobot()
    fin = False

    rospy.Subscriber('program_started', data_employee, nav.callback)
    rospy.Subscriber('/recognition', data_recognition, nav.get_recog)
    rospy.Subscriber('/mobile_base/events/bumper', BumperEvent, nav.collect)

    while(fin == False and not rospy.is_shutdown()):
        if(st == True):

            for i in range(1, tam):
                nav.movebase_client(i)
                nav.spin_move(i)
                
            
            if(nav.collected == False):
                nav.movebase_client(0)

                TOKEN = "6091287301:AAEOOX5aOx_BeWiGsToPHNRMlIajqAnRyXE"
                chat_id = str(nav.id_tele)
                #print(chat_id)

                message = "Tienes un paquete en el punto base. Por favor, vaya a recogerlo."
                url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat_id}&text={message}"
                print(requests.get(url).json()) # this sends the message
            
            else:
                nav.movebase_client(-1)

            fin = True

