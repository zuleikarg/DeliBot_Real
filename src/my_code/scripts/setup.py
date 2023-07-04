#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl
import cv2 as cv
import dlib
import numpy as np
import math
import imutils

import os

import rospy


class InterEmplo(Frame):
 
    def __init__(self, master, *args, **kwargs):
        Frame.__init__(self, master, *args, **kwargs)
        self.folder = os.environ.get('DELIBOT_PATH', '')

        self.parent = master
        self.grid()
        self.once = 0
        self.createWidgets()

    def finish(self, new_image, selection):
        
        cv.imwrite(self.folder + "/DeliBot_Real/src/siam-mot/demos/conocidos/"+selection +".jpg", new_image)

        os.remove(self.folder + "/DeliBot_Real/src/siam-mot/demos/"+selection +".jpg")
        
        self.quit()

    def show_selection(self):
        # Obtener la opción seleccionada.
        selection = self.combo.get()

        if(selection != ''):
            
            cam = cv.VideoCapture(0)
            result = 0
            while(result == 0):
                result, image = cam.read()

            cam.release()

            image = cv.flip(image,1)

            detector = dlib.get_frontal_face_detector()
            predictor = dlib.shape_predictor(self.folder + "/DeliBot_Real/src/my_code/dataset/shape_predictor_68_face_landmarks.dat")

            gray = cv.cvtColor(image, cv.COLOR_BGR2GRAY)
            # Detect the face
            rects = detector(gray, 1)
            # Detect landmarks for each face
            shape = np.zeros((68, 2), dtype="int")

            for rect in rects:
                # Get the landmark points
                shape = predictor(gray, rect)
            # Convert it to the NumPy Array
                shape_np = np.zeros((68, 2), dtype="int")
                for i in range(0, 68):
                    shape_np[i] = (shape.part(i).x, shape.part(i).y)
                shape = shape_np
            
            modified = False

            rotated_image = image
            if(not shape[0][0] == 0 or not shape[67][0] == 0):
                diference = shape_np[45] - shape_np[36]
                # print(diference)
                angle = -(math.atan2(diference[1], diference[0])*180)/math.pi   #HABRA QUE EXPLICAR EL PORQUE DE ESE SIGNO -
                
                # print(angle)
                rotated_image = imutils.rotate_bound(image,angle)

                modified = True

            if(modified):
                gray = cv.cvtColor(rotated_image, cv.COLOR_BGR2GRAY)
                # Detect the face
                rects = detector(gray, 1)
                # Detect landmarks for each face

                for rect in rects:
                    # Get the landmark points
                    shape2 = predictor(gray, rect)
                # Convert it to the NumPy Array
                    shape_np2 = np.zeros((68, 2), dtype="int")
                    for i in range(0, 68):
                        shape_np2[i] = (shape2.part(i).x, shape2.part(i).y)
                    shape = shape_np2

            new_image = rotated_image[shape[19][1]:shape[8][1],shape[0][0]:shape[16][0]]

            if(new_image.size != 0):

                cv.imwrite(self.folder + "/DeliBot_Real/src/siam-mot/demos/"+selection +".jpg", new_image)

                disp = "Se ha seleccionado al empleado "+selection

                self.display2 = Text(self.p2, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
                self.display2.insert(INSERT,disp)
                self.display2.place(x=20,y=200)
                self.display2.config(state='disabled')

                # Define variable to load the dataframe
                dataframe = openpyxl.load_workbook(self.folder + "/DeliBot_Real/src/my_code/datos_empleados.xlsx")
                # Define variable to read sheet
                dataframe1 = dataframe.active
                
                # Iterate the loop to read the cell values
                for row in dataframe1.iter_rows(1,dataframe1.max_row):
                    if(row[0].value == selection):
                        break

                self.display2 = Text(self.p2, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
                self.display2.insert(INSERT,row[1].value)
                self.display2.place(x=20,y=250)
                self.display2.config(state='disabled')

                # Create an object of tkinter ImageTk
                self.img = Image.open(self.folder + "/DeliBot_Real/src/siam-mot/demos/"+selection +".jpg")

                self.img = ImageTk.PhotoImage(self.img.resize((200,200), Image.Resampling.LANCZOS))

                if(self.once == 0):
                    self.once = 1
                else:
                    self.label.after(100, self.label.destroy())

                # Create a Label Widget to display the text or Image
                self.label = Label(self.p2, image = self.img)
                # self.label.place(relx=200,rely=200)
                self.label.pack()

                self.ceButton = Button(self.p2, font=("Arial", 10), fg='black', text="Guardar", highlightbackground='white', command=lambda: self.finish(new_image, selection))
                self.ceButton.place(x=220,y=300)

                self.notebook.tab(self.p2, state='normal')

            else:
                print("Cara no detectada")

    def createWidgets(self):
        # self.display.grid(row=0, column=0, columnspan=4, sticky="nsew")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand='yes')
        self.p1 = ttk.Frame(self.notebook,width=500, height=500)
        self.p2 = ttk.Frame(self.notebook,width=500, height=500)
        self.p1.pack(fill='both', expand=True)
        self.p1.pack(fill='both', expand=True)

        # Define variable to load the dataframe
        dataframe = openpyxl.load_workbook(self.folder + "/DeliBot_Real/src/my_code/datos_empleados.xlsx")
        # Define variable to read sheet
        dataframe1 = dataframe.active
        
        names = []
        # Iterate the loop to read the cell values
        for row in dataframe1.iter_rows(2,dataframe1.max_row):
            names.append(row[0].value)

        self.combo = ttk.Combobox(self.p1, width= 30, height= 10, state="readonly",values=names)
        self.combo.place(x=140,y=175)

        self.display1 = Text(self.p1, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
        self.display1.insert(INSERT,"Esta es la interfaz para la introducción de un empleado")
        self.display1.place(x=20,y=50)
        self.display1.config(state='disabled')

        # self.display.insert(1, "Est-a es la interfaz para la selección del empleado a recibir el paquete")       
        self.notebook.add(self.p1,text='Toma de foto')
        self.notebook.add(self.p2,text='Info empleado',state='disabled')

        self.ceButton = Button(self.p1, font=("Arial", 10), fg='black', text="Tomar foto", highlightbackground='white', command=lambda: self.show_selection())
        self.ceButton.place(x=180,y=350)

 
if __name__ == "__main__":
    rospy.init_node('setup', anonymous=True)
    
    Interface = Tk()
    Interface.title("Interfaz de introducción de datos de empleado")
    Interface.resizable(False, False)
    Interface.geometry('500x500')
    # Interface.config(cursor="pencil")
    root = InterEmplo(Interface).grid()

    Interface.mainloop()