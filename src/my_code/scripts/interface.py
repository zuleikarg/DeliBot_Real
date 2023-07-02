#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl
import os

import rospy
from my_code.msg import data_employee

name = ''
depart = ''

class InterEmplo(Frame):
 
    def __init__(self, master, *args, **kwargs):
        Frame.__init__(self, master, *args, **kwargs)
        self.folder = os.environ.get('DELIBOT_PATH', '')

        self.parent = master
        self.grid()
        self.once = 0
        self.createWidgets()
    
    def get_data(self):

        return self.selection, self.dep

    def send(self):
        self.quit()

    def show_selection(self):
        global name, depart
        # Obtener la opción seleccionada.
        selection = self.combo.get()
        name = selection

        if(selection != ''):
            disp = "Se ha seleccionado al empleado "+selection

            self.display2 = Text(self.p2, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
            self.display2.insert(INSERT,disp)
            self.display2.place(x=20,y=200)
            self.display2.config(state='disabled')

            # Define variable to load the dataframe
            dataframe = openpyxl.load_workbook(self.folder + "/tfg_ros_real/src/my_code/datos_empleados.xlsx")
            # Define variable to read sheet
            dataframe1 = dataframe.active
            
            # Iterate the loop to read the cell values
            for row in dataframe1.iter_rows(1,dataframe1.max_row):
                if(row[0].value == selection):
                    break
            
            depart = row[1].value

            self.display2 = Text(self.p2, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
            self.display2.insert(INSERT,row[1].value)
            self.display2.place(x=20,y=250)
            self.display2.config(state='disabled')

            # Create an object of tkinter ImageTk
            self.img = Image.open(self.folder + "/tfg_ros_real/src/my_code/fotos/"+row[2].value +".png")

            self.img = ImageTk.PhotoImage(self.img.resize((300,150), Image.Resampling.LANCZOS))

            if(self.once == 0):
                self.once = 1
            else:
                self.label.after(100, self.label.destroy())

            # Create a Label Widget to display the text or Image
            self.label = Label(self.p2, image = self.img)
            # self.label.place(relx=200,rely=200)
            self.label.pack()

            self.ceButton = Button(self.p2, font=("Arial", 10), fg='black', text="Enviar", highlightbackground='white', command=lambda: self.send())
            self.ceButton.place(x=220,y=300)

            self.notebook.tab(self.p2, state='normal') 

    def createWidgets(self):
        # self.display.grid(row=0, column=0, columnspan=4, sticky="nsew")

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand='yes')
        self.p1 = ttk.Frame(self.notebook,width=500, height=500)
        self.p2 = ttk.Frame(self.notebook,width=500, height=500)
        self.p1.pack(fill='both', expand=True)
        self.p1.pack(fill='both', expand=True)


        # Define variable to load the dataframe
        dataframe = openpyxl.load_workbook(self.folder + "/tfg_ros_real/src/my_code/datos_empleados.xlsx")
        # Define variable to read sheet
        dataframe1 = dataframe.active
        
        names = []
        # Iterate the loop to read the cell values
        for row in dataframe1.iter_rows(2,dataframe1.max_row):
            names.append(row[0].value)

        self.combo = ttk.Combobox(self.p1, width= 30, height= 10, state="readonly",values=names)
        self.combo.place(x=140,y=175)

        self.display1 = Text(self.p1, width= 56, height= 1,font=("Arial", 11), relief=RAISED, bg='white', fg='black')
        self.display1.insert(INSERT,"Esta es la interfaz para la selección del empleado a recibir el paquete")
        self.display1.place(x=20,y=50)
        self.display1.config(state='disabled')

        # self.display.insert(1, "Est-a es la interfaz para la selección del empleado a recibir el paquete")       
        self.notebook.add(self.p1,text='Selcción empleado')
        self.notebook.add(self.p2,text='Info empleado',state='disabled')

        self.ceButton = Button(self.p1, font=("Arial", 10), fg='black', text="Insertar respuesta", highlightbackground='white', command=lambda: self.show_selection())
        self.ceButton.place(x=180,y=350)

 
if __name__ == "__main__":
    rospy.init_node('interface', anonymous=True)
    pub = rospy.Publisher('program_started', data_employee, queue_size=10)
    
    Interface = Tk()
    Interface.title("Interfaz de elección de empleado")
    Interface.resizable(False, False)
    Interface.geometry('500x500')
    # Interface.config(cursor="pencil")
    root = InterEmplo(Interface).grid()

    Interface.mainloop()

    if(name != '' and depart != ''):
        message = data_employee()
        message.name = name
        message.department = depart

        pub.publish(message)
        print(f"Se ha seleccionado {name} del {depart}")